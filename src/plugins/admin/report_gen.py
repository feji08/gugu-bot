from io import BytesIO
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ...database import engine


def generate_report_xlsx(start_date: datetime, end_date: datetime) -> bytes:
    query = """
        SELECT
            users.user_id,
            users.nickname,
            checkin_records.checkin_time,
            assignments.name AS assignment_name
        FROM users
        LEFT JOIN checkin_records ON checkin_records.user_id = users.user_id
        LEFT JOIN assignments ON checkin_records.assignment_id = assignments.id
    """
    df = pd.read_sql_query(query, engine)
    df["checkin_time"] = pd.to_datetime(df["checkin_time"], format="mixed")

    dates = pd.date_range(start_date, end_date)
    user_nicknames = df["nickname"].unique()
    result_df = pd.DataFrame(index=user_nicknames, columns=dates.date)

    for nickname in user_nicknames:
        user_records = df[df["nickname"] == nickname]
        for date in dates:
            start_time = datetime.combine(date, datetime.min.time())
            end_time = start_time + timedelta(days=1)

            checkin_record = user_records[
                (user_records["checkin_time"] >= start_time)
                & (user_records["checkin_time"] < end_time)
            ]

            date_record = ""
            if not checkin_record.empty:
                name = checkin_record.iloc[0]["assignment_name"]
                if "练笔" in str(name):
                    date_record = "输出练笔"
                elif "扒文" in str(name):
                    date_record = "扒文扒榜"
                elif "摘抄" in str(name):
                    date_record = "摘抄作业"
                elif "节奏" in str(name):
                    date_record = "节奏练习"
                elif "请假" in str(name):
                    date_record = "请假"
                else:
                    date_record = "其他练习"

            result_df.loc[nickname, date.date()] = date_record

    # --- 统计 ---
    stats_rows = {}
    for nickname in user_nicknames:
        row = result_df.loc[nickname]
        values = [str(v) for v in row.values]

        leave_count = sum(1 for v in values if v == "请假")
        checkin_count = sum(1 for v in values if v and v != "请假")
        practice_count = sum(1 for v in values if v == "输出练笔")
        review_count = sum(1 for v in values if v == "扒文扒榜")

        reward = 0
        streak = 0
        for v in values:
            if v and v != "请假":
                streak += 1
            else:
                reward += streak // 7
                streak = 0
        reward += streak // 7

        stats_rows[nickname] = {
            "打卡次数": checkin_count,
            "请假次数": leave_count,
            "规则奖励": reward,
            "输出练笔": practice_count,
            "扒文扒榜": review_count,
        }

    # 转置：行=日期，列=昵称
    result_df = result_df.T
    result_df.index.name = "日期"

    # 追加统计行
    stats_df = pd.DataFrame(stats_rows)
    stats_df.index.name = "日期"
    result_df = pd.concat([result_df, stats_df])

    # --- 写入 xlsx 并渲染样式 ---
    buf = BytesIO()
    result_df.to_excel(buf, engine="openpyxl")
    buf.seek(0)

    from openpyxl import load_workbook
    wb = load_workbook(buf)
    ws = wb.active

    # 样式定义
    green_fill = PatternFill(start_color="B3D8A8", end_color="B3D8A8", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE5D9", end_color="FFE5D9", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    num_cols = len(user_nicknames) + 1  # A列(日期) + 昵称列
    num_date_rows = len(dates)
    stats_labels = ["打卡次数", "请假次数", "规则奖励", "输出练笔", "扒文扒榜"]
    num_stats = len(stats_labels)

    # 插入大表头行（第1行上方插入一行）
    ws.insert_rows(1)
    title = f"{end_date.year}年{end_date.month}月打卡统计（{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}）"
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=1, column=1).fill = gray_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    # 合并后每个单元格都要设灰底才生效
    for col in range(2, num_cols + 1):
        ws.cell(row=1, column=col).fill = gray_fill

    # 表头行（第2行）加粗居中灰底
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = bold_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = gray_fill
    ws.cell(row=2, column=1, value="日期/昵称")

    # 数据区域：第3行开始，日期行 + 统计行
    for row_idx in range(3, 3 + num_date_rows + num_stats):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = center
            val = str(cell.value or "")

            # 日期数据行涂色
            if row_idx < 3 + num_date_rows and col_idx > 1:
                if val == "请假":
                    cell.fill = orange_fill
                elif val and val != "None":
                    cell.fill = green_fill

            # 统计行样式
            if row_idx >= 3 + num_date_rows:
                if col_idx == 1:
                    cell.fill = gray_fill
                cell.font = bold_font

    # 列宽自适应
    ws.column_dimensions["A"].width = 14
    for col in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    buf2 = BytesIO()
    wb.save(buf2)
    return buf2.getvalue()
